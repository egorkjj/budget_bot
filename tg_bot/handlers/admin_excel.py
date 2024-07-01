from openpyxl import Workbook as WB
from openpyxl.styles import Font, Alignment
from aiogram import Dispatcher, types
from aiogram.dispatcher import FSMContext
from aiogram.types import InputFile
from tg_bot.DBSM import review_for_day_for_managers, month_rev
import os
from datetime import datetime
import pytz

month_names_ru = {
    1: 'Январь',
    2: 'Февраль',
    3: 'Март',
    4: 'Апрель',
    5: 'Май',
    6: 'Июнь',
    7: 'Июль',
    8: 'Август',
    9: 'Сентябрь',
    10: 'Октябрь',
    11: 'Ноябрь',
    12: 'Декабрь'
}


def register_excel(dp: Dispatcher):
    dp.register_message_handler(cash_reviews_for_today, commands = ["cash"])
    dp.register_message_handler(month_excel, commands = ["month"])


async def cash_reviews_for_today(message: types.Message, state: FSMContext):
    data = review_for_day_for_managers()
    result = data[0]
    cash_res = data[1]
    
    wb = WB()
    wb.remove(wb["Sheet"])
    sheet1 = wb.create_sheet("Виталий", 0)
    sheet2 = wb.create_sheet("Наталья", 1)
    sheet3 = wb.create_sheet("Екатерина", 2)
    sheets = [sheet1, sheet2, sheet3]

    for sheet in sheets:
        sheet["A1"] = "Доход/расход"
        sheet['A1'].font = Font(color="FF0000")
        sheet["B1"] = "Категория"
        sheet['B1'].font = Font(color="FF0000")
        sheet["C1"] = "Комментарий"
        sheet['C1'].font = Font(color="FF0000")
        sheet["D1"] = "Сумма"
        sheet['D1'].font = Font(color="FF0000") 
        sheet["E1"] = "ФИО"
        sheet['E1'].font = Font(color="FF0000") 

    for i in range(3):
        shit = sheets[i]
        query = result[i]
        cash = cash_res[i]
        for j in range(len(query)):
            shit[f"A{j+2}"] = "Доход" if query[j]["is_income"] else "Расход"
            shit[f"B{j+2}"] = query[j]["type"] if query[j]["type"] != "отсутствует" else "-"
            shit[f"C{j+2}"] = query[j]["comment"] if query[j]["comment"] != "отсутствует" else "❌"
            shit[f"D{j+2}"] = f'{query[j]["summ"]} бел.руб.'
            shit[f"E{j+2}"] = query[j]["fio"]


        row = len(query) + 2
        shit.merge_cells(f"B{row}:C{row}")
        shit.merge_cells(f"B{row+1}:C{row+1}")
        shit[f"B{row+1}"] = f"{cash} бел.руб."
        shit[f"B{row}"] = "Итого"
        shit[f"B{row}"].font = Font(color = "FF0000")
        shit[f"B{row+1}"].font = Font(color = "FF0000")
        shit[f"B{row}"].alignment = Alignment(horizontal= "center")
        shit[f"B{row+1}"].alignment = Alignment(horizontal= "center")

    if not os.path.isdir("tg_bot/excel"):
        os.mkdir("tg_bot/excel")
    date = datetime.strftime(datetime.now(pytz.timezone('Europe/Moscow')), "%d.%m.%Y")
    wb.save(f"tg_bot/excel/Касса {date}.xlsx")

    await message.answer_document(document = InputFile(f"tg_bot/excel/Касса {date}.xlsx"))
    os.remove(f"tg_bot/excel/Касса {date}.xlsx")


async def month_excel(message: types.Message, state: FSMContext):
    data = month_rev()
    
    wb = WB()
    wb.remove(wb["Sheet"])
    sheet1 = wb.create_sheet("Доходы", 0)
    sheet2 = wb.create_sheet("Расходы", 0)
    sheets = [sheet1, sheet2]

    for sheet in sheets:
        sheet["A1"] = "Менеджер"
        sheet['A1'].font = Font(color="FF0000")
        sheet["B1"] = "Тип"
        sheet['B1'].font = Font(color="FF0000")
        sheet["C1"] = "Дата"
        sheet['C1'].font = Font(color="FF0000")
        sheet["D1"] = "Категория"
        sheet['D1'].font = Font(color="FF0000")
        sheet["E1"] = "Комментарий"
        sheet['E1'].font = Font(color="FF0000")
        sheet["F1"] = "Сумма"
        sheet['F1'].font = Font(color="FF0000")
        sheet["G1"] = "Оплата"
        sheet['G1'].font = Font(color="FF0000")
        sheet["H1"] = "ФИО"
        sheet['H1'].font = Font(color="FF0000")

     
    for i in range(2):
        shit = sheets[i]
        query = data[0][i]
        sums = data[1][i]
        for j in range(len(query)):
            shit[f"A{j+2}"] = query[j]["manager"]
            shit[f"B{j+2}"] = query[j]["type"]
            shit[f"C{j+2}"] = query[j]["date"]
            shit[f"D{j+2}"] = query[j]["cat"]
            shit[f"E{j+2}"] = query[j]["comment"]
            shit[f"F{j+2}"] = query[j]["summ"]
            shit[f"G{j+2}"] = query[j]["pay"]
            shit[f"H{j+2}"] = query[j]["fio"]



        row = len(query) + 2
        shit.merge_cells(f"D{row}:E{row}")
        shit.merge_cells(f"D{row+1}:E{row+1}")
        shit[f"D{row+1}"] = f"{sums} бел.руб."
        shit[f"D{row}"] = "Доходы за месяц" if i == 0 else "Расходы за месяц"
        shit[f"D{row}"].font = Font(color = "FF0000")
        shit[f"D{row+1}"].font = Font(color = "FF0000")
        shit[f"D{row}"].alignment = Alignment(horizontal= "center")
        shit[f"D{row+1}"].alignment = Alignment(horizontal= "center")
    
    if not os.path.isdir("tg_bot/excel"):
        os.mkdir("tg_bot/excel")
    month = datetime.now().month
    month = month_names_ru[month]
    year = datetime.now().year
    wb.save(f"tg_bot/excel/Отчет за {month} {year}.xlsx")

    await message.answer_document(document = InputFile(f"tg_bot/excel/Отчет за {month} {year}.xlsx"))
    os.remove(f"tg_bot/excel/Отчет за {month} {year}.xlsx")



